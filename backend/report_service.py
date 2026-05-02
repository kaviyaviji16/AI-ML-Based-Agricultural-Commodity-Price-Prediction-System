"""
Report Generation Service
Generates PDF and Excel reports with real price data and predictions
"""
import os
import io
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc
from api.models.database import RawPrice, Prediction, Recommendation

REPORTS_DIR = Path("reports")
REPORTS_DIR.mkdir(exist_ok=True)


async def get_price_data(db: AsyncSession, commodities: list, days: int) -> pd.DataFrame:
    cutoff = datetime.utcnow() - timedelta(days=days)
    result = await db.execute(
        select(RawPrice)
        .where(RawPrice.date >= cutoff)
        .order_by(RawPrice.date)
    )
    rows = result.scalars().all()
    if not rows:
        return pd.DataFrame()
    data = []
    for r in rows:
        comm = r.commodity if isinstance(r.commodity, str) else r.commodity.value
        if not commodities or comm in commodities:
            data.append({
                'commodity': comm,
                'date': r.date,
                'market': r.market,
                'state': r.state,
                'modal_price': r.modal_price,
                'min_price': r.min_price,
                'max_price': r.max_price,
                'arrivals_tonnes': r.arrivals_tonnes,
            })
    return pd.DataFrame(data)


async def get_predictions_data(db: AsyncSession, commodities: list) -> pd.DataFrame:
    result = await db.execute(
        select(Prediction).order_by(desc(Prediction.created_at)).limit(200)
    )
    rows = result.scalars().all()
    data = []
    for r in rows:
        comm = r.commodity if isinstance(r.commodity, str) else r.commodity.value
        if not commodities or comm in commodities:
            expl = r.explanation or {}
            data.append({
                'commodity': comm,
                'horizon_days': r.horizon_days,
                'predicted_price': r.predicted_price,
                'confidence_score': r.confidence_score,
                'price_change_pct': expl.get('price_change_pct', 0),
                'scenario': r.scenario,
                'target_date': r.target_date,
            })
    return pd.DataFrame(data)


async def get_recommendations_data(db: AsyncSession) -> pd.DataFrame:
    result = await db.execute(
        select(Recommendation).order_by(desc(Recommendation.generated_at)).limit(50)
    )
    rows = result.scalars().all()
    data = []
    for r in rows:
        comm = r.commodity if isinstance(r.commodity, str) else r.commodity.value
        risk = r.risk_level if isinstance(r.risk_level, str) else r.risk_level.value
        data.append({
            'commodity': comm,
            'headline': r.headline,
            'action_type': r.action_type,
            'quantity_tonnes': r.quantity_tonnes,
            'risk_level': risk,
            'confidence_score': r.confidence_score,
            'status': r.status,
            'generated_at': r.generated_at,
        })
    return pd.DataFrame(data)


async def generate_excel_report(
    db: AsyncSession,
    report_type: str,
    commodities: list,
    start_date=None,
    end_date=None,
) -> str:
    """Generate Excel report and return file path"""

    # Determine date range
    days_map = {
        'weekly_summary': 7,
        'monthly_analysis': 30,
        'annual_review': 365,
        'custom': 30,
    }
    days = days_map.get(report_type, 30)

    # Get data
    prices_df = await get_price_data(db, commodities, days)
    preds_df = await get_predictions_data(db, commodities)
    recs_df = await get_recommendations_data(db)

    # Create Excel file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"agri_report_{report_type}_{timestamp}.xlsx"
    filepath = REPORTS_DIR / filename

    with pd.ExcelWriter(str(filepath), engine='openpyxl') as writer:
        # Sheet 1: Summary
        summary_data = {
            'Report Type': [report_type.replace('_', ' ').title()],
            'Generated At': [datetime.now().strftime('%d %b %Y %I:%M %p')],
            'Date Range': [f"Last {days} days"],
            'Commodities': [', '.join(commodities) if commodities else 'All 8 Commodities'],
            'Total Price Records': [len(prices_df)],
            'Total Predictions': [len(preds_df)],
            'Total Recommendations': [len(recs_df)],
        }
        pd.DataFrame(summary_data).T.reset_index().rename(
            columns={'index': 'Parameter', 0: 'Value'}
        ).to_excel(writer, sheet_name='Summary', index=False)

        # Sheet 2: Price Data
        if not prices_df.empty:
            # Pivot: average price per commodity per day
            price_pivot = prices_df.groupby(['commodity', 'date'])['modal_price'].mean().reset_index()
            price_pivot['date'] = pd.to_datetime(price_pivot['date']).dt.strftime('%d-%b-%Y')
            price_pivot.columns = ['Commodity', 'Date', 'Avg Modal Price (Rs/kg)']
            price_pivot.to_excel(writer, sheet_name='Price_Data', index=False)

            # Sheet 3: Market wise prices
            market_prices = prices_df.groupby(['commodity', 'market'])['modal_price'].agg(
                ['mean', 'min', 'max']
            ).reset_index()
            market_prices.columns = ['Commodity', 'Market', 'Avg Price', 'Min Price', 'Max Price']
            market_prices = market_prices.round(2)
            market_prices.to_excel(writer, sheet_name='Market_Prices', index=False)

        # Sheet 4: Predictions
        if not preds_df.empty:
            preds_display = preds_df.copy()
            preds_display['target_date'] = pd.to_datetime(
                preds_display['target_date']
            ).dt.strftime('%d-%b-%Y')
            preds_display.columns = [
                'Commodity', 'Horizon (Days)', 'Predicted Price (Rs/kg)',
                'Confidence (%)', 'Price Change (%)', 'Scenario', 'Target Date'
            ]
            preds_display = preds_display.round(2)
            preds_display.to_excel(writer, sheet_name='Predictions', index=False)

        # Sheet 5: Recommendations
        if not recs_df.empty:
            recs_display = recs_df.copy()
            recs_display['generated_at'] = pd.to_datetime(
                recs_display['generated_at']
            ).dt.strftime('%d-%b-%Y %I:%M %p')
            recs_display.columns = [
                'Commodity', 'Recommendation', 'Action Type',
                'Quantity (Tonnes)', 'Risk Level', 'Confidence (%)',
                'Status', 'Generated At'
            ]
            recs_display.to_excel(writer, sheet_name='Recommendations', index=False)

        # Sheet 6: Commodity Stats
        if not prices_df.empty:
            stats = prices_df.groupby('commodity')['modal_price'].agg([
                'mean', 'min', 'max', 'std', 'count'
            ]).reset_index()
            stats.columns = [
                'Commodity', 'Avg Price (Rs/kg)', 'Min Price',
                'Max Price', 'Std Dev', 'Data Points'
            ]
            stats = stats.round(2)
            stats.to_excel(writer, sheet_name='Commodity_Stats', index=False)

        # Style the workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = writer.book
        header_fill = PatternFill(start_color="1a5c38", end_color="1a5c38", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Style header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    return str(filepath)


async def generate_pdf_report(
    db: AsyncSession,
    report_type: str,
    commodities: list,
    start_date=None,
    end_date=None,
) -> str:
    """Generate PDF report and return file path"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.units import cm
    from reportlab.lib import colors

    days_map = {
        'weekly_summary': 7,
        'monthly_analysis': 30,
        'annual_review': 365,
        'custom': 30,
    }
    days = days_map.get(report_type, 30)

    prices_df = await get_price_data(db, commodities, days)
    preds_df = await get_predictions_data(db, commodities)
    recs_df = await get_recommendations_data(db)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"agri_report_{report_type}_{timestamp}.pdf"
    filepath = REPORTS_DIR / filename

    doc = SimpleDocTemplate(str(filepath), pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    GREEN = HexColor('#1a5c38')
    LIGHT_GREEN = HexColor('#d4f0e0')
    DARK = HexColor('#111827')

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', fontSize=20, textColor=GREEN,
                                  fontName='Helvetica-Bold', spaceAfter=6)
    subtitle_style = ParagraphStyle('Subtitle', fontSize=11, textColor=DARK,
                                     fontName='Helvetica', spaceAfter=12)
    heading_style = ParagraphStyle('Heading', fontSize=13, textColor=GREEN,
                                    fontName='Helvetica-Bold', spaceAfter=8, spaceBefore=16)
    normal_style = ParagraphStyle('Normal2', fontSize=9, fontName='Helvetica',
                                   spaceAfter=4, textColor=DARK)

    story = []

    # ── Header ──────────────────────────────────────────────────────────────
    story.append(Paragraph("🌾 AgriPrice Intelligence System", title_style))
    story.append(Paragraph(
        f"Ministry of Consumer Affairs, Food & Public Distribution — Government of India",
        subtitle_style
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=GREEN))
    story.append(Spacer(1, 0.3*cm))

    report_titles = {
        'weekly_summary': 'Weekly Price Summary Report',
        'monthly_analysis': 'Monthly Price Analysis Report',
        'annual_review': 'Annual Review Report',
        'custom': 'Custom Analysis Report',
    }
    story.append(Paragraph(report_titles.get(report_type, 'Price Report'), heading_style))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')} | "
        f"Period: Last {days} days | "
        f"Commodities: {', '.join(commodities) if commodities else 'All 8'}",
        normal_style
    ))
    story.append(Spacer(1, 0.5*cm))

    # ── Executive Summary ────────────────────────────────────────────────────
    story.append(Paragraph("Executive Summary", heading_style))
    summary_data = [
        ['Parameter', 'Value'],
        ['Report Period', f'Last {days} days'],
        ['Commodities Covered', ', '.join(commodities) if commodities else 'All 8 commodities'],
        ['Price Records Analyzed', str(len(prices_df))],
        ['Predictions Generated', str(len(preds_df))],
        ['Active Recommendations', str(len(recs_df))],
        ['High Risk Alerts', str(len(recs_df[recs_df['risk_level']=='high'])) if not recs_df.empty else '0'],
    ]
    t = Table(summary_data, colWidths=[8*cm, 9*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GREEN),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [LIGHT_GREEN, white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # ── Current Prices ───────────────────────────────────────────────────────
    if not prices_df.empty:
        story.append(Paragraph("Current Commodity Prices (Rs/kg)", heading_style))
        stats = prices_df.groupby('commodity')['modal_price'].agg(
            ['mean', 'min', 'max']
        ).reset_index().round(2)

        price_table_data = [['Commodity', 'Avg Price', 'Min Price', 'Max Price', 'Trend']]
        for _, row in stats.iterrows():
            comm = str(row['commodity']).title()
            avg = f"Rs.{row['mean']:.2f}"
            mn = f"Rs.{row['min']:.2f}"
            mx = f"Rs.{row['max']:.2f}"
            spread = row['max'] - row['min']
            trend = '▲ Volatile' if spread > row['mean'] * 0.3 else '→ Stable'
            price_table_data.append([comm, avg, mn, mx, trend])

        pt = Table(price_table_data, colWidths=[4*cm, 3*cm, 3*cm, 3*cm, 4*cm])
        pt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), GREEN),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [LIGHT_GREEN, white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(pt)
        story.append(Spacer(1, 0.5*cm))

    # ── Price Predictions ────────────────────────────────────────────────────
    if not preds_df.empty:
        story.append(Paragraph("AI Price Predictions", heading_style))
        pred_table_data = [['Commodity', 'Horizon', 'Predicted Price', 'Change %', 'Confidence', 'Scenario']]
        for _, row in preds_df.head(20).iterrows():
            pct = row['price_change_pct']
            arrow = '▲' if pct > 0 else '▼'
            pred_table_data.append([
                str(row['commodity']).title(),
                f"{row['horizon_days']}d",
                f"Rs.{row['predicted_price']:.2f}",
                f"{arrow} {abs(pct):.1f}%",
                f"{row['confidence_score']:.0f}%",
                str(row['scenario']).title(),
            ])

        pred_t = Table(pred_table_data, colWidths=[3.5*cm, 2*cm, 3.5*cm, 2.5*cm, 2.5*cm, 3*cm])
        pred_t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), GREEN),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [LIGHT_GREEN, white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('PADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(pred_t)
        story.append(Spacer(1, 0.5*cm))

    # ── Recommendations ──────────────────────────────────────────────────────
    if not recs_df.empty:
        story.append(Paragraph("Buffer Stock Recommendations", heading_style))
        rec_table_data = [['Commodity', 'Action', 'Risk', 'Quantity (T)', 'Confidence', 'Status']]
        for _, row in recs_df.head(10).iterrows():
            rec_table_data.append([
                str(row['commodity']).title(),
                str(row['action_type']).replace('_', ' ').title(),
                str(row['risk_level']).upper(),
                f"{row['quantity_tonnes']:.0f}" if row['quantity_tonnes'] else '—',
                f"{row['confidence_score']:.0f}%" if row['confidence_score'] else '—',
                str(row['status']).title(),
            ])

        rec_t = Table(rec_table_data, colWidths=[3*cm, 4*cm, 2*cm, 3*cm, 3*cm, 2*cm])
        rec_t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), GREEN),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [LIGHT_GREEN, white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('PADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(rec_t)

    # ── Footer ───────────────────────────────────────────────────────────────
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=GREEN))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        "This report was generated automatically by the AgriPrice Intelligence System. "
        "For official use only — Ministry of Consumer Affairs, Government of India.",
        ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, fontName='Helvetica-Oblique')
    ))

    doc.build(story)
    return str(filepath)